"""
excel_report.py — Professional Excel workbook export.

Single-page mode   → save_excel(result, path)
Site-crawl mode    → save_site_excel(crawl_result, path)

Workbook sheets:
  1. Summary          — score dashboard, category scores table
  2. All Findings     — every finding across all pages, filterable
  3. Quick Wins       — severity CRITICAL/WARNING with effort=Quick Win
  4. Roadmap          — Phase 1 / 2 / 3 action items
  5. Checklist        — site-type-specific custom checks
  6. Traffic Strategy — profile traffic notes
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from .models import AuditResult
    from ..site_crawler import SiteCrawlResult

from ..categories.base import Severity


# ─── Colour palette (openpyxl PatternFill hex strings) ────────────────────────

_COL = {
    # Severity fills
    "critical_fill":  "FFCCCC",   # soft red
    "warning_fill":   "FFF3CD",   # soft amber
    "info_fill":      "CCE5FF",   # soft blue
    "pass_fill":      "D4EDDA",   # soft green
    "ext_fill":       "E8D5F5",   # soft purple (external tool needed)
    # Header fills
    "header_dark":    "1A202C",   # near-black
    "header_mid":     "2D3748",   # dark grey
    "header_phase1":  "C53030",   # red (Phase 1)
    "header_phase2":  "B7791F",   # amber (Phase 2)
    "header_phase3":  "2B6CB0",   # blue (Phase 3)
    "header_green":   "276749",   # green (traffic)
    # Score bands
    "score_low":      "E53E3E",
    "score_mid":      "D69E2E",
    "score_high":     "38A169",
    # Neutral
    "white":          "FFFFFF",
    "light_grey":     "F7FAFC",
    "mid_grey":       "EDF2F7",
}

_SEV_FILL = {
    Severity.CRITICAL: _COL["critical_fill"],
    Severity.WARNING:  _COL["warning_fill"],
    Severity.INFO:     _COL["info_fill"],
    Severity.PASS:     _COL["pass_fill"],
}


def _try_import():
    try:
        import openpyxl
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
        return openpyxl, PatternFill, Font, Alignment, Border, Side, get_column_letter
    except ImportError:
        return None, None, None, None, None, None, None


def _fill(hex_color: str):
    from openpyxl.styles import PatternFill
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=11, italic=False):
    from openpyxl.styles import Font
    return Font(bold=bold, color=color, size=size, italic=italic)


def _align(h="left", v="top", wrap=True):
    from openpyxl.styles import Alignment
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border_thin():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _score_color(score: int) -> str:
    if score >= 70:
        return _COL["score_high"]
    if score >= 45:
        return _COL["score_mid"]
    return _COL["score_low"]


def _set_col_widths(ws, widths: list[int]) -> None:
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, values: list[str], fill_hex: str, row: int = 1) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(fill_hex)
        cell.font = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = _align(h="center", v="center", wrap=False)
        cell.border = _border_thin()


def _data_cell(ws, row: int, col: int, value, fill_hex: str = "FFFFFF",
               bold: bool = False, color: str = "2D3748") -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(fill_hex)
    cell.font = _font(bold=bold, color=color, size=10)
    cell.alignment = _align()
    cell.border = _border_thin()


# ─── Sheet builders ───────────────────────────────────────────────────────────

def _sheet_summary_single(wb, result: "AuditResult") -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:F1")
    ws["A1"] = "SEO AUDIT REPORT"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = _fill(_COL["header_dark"])
    ws["A1"].alignment = _align(h="center", v="center", wrap=False)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ws["A2"] = result.url
    ws["A2"].font = _font(italic=True, color="718096", size=10)
    ws["A2"].fill = _fill(_COL["light_grey"])
    ws["A2"].alignment = _align(h="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = f"Site type: {result.profile.label}   |   Generated: {datetime.now().strftime('%d %b %Y')}"
    ws["A3"].font = _font(color="718096", size=9)
    ws["A3"].fill = _fill(_COL["mid_grey"])
    ws["A3"].alignment = _align(h="center")

    # Score block
    row = 5
    score_labels = ["Overall Score", "Critical Issues", "Warnings", "Quick Wins"]
    score_values = [
        f"{result.overall_score}/100",
        str(len(result.critical_findings)),
        str(len(result.warning_findings)),
        str(len(result.quick_wins)),
    ]
    score_colors = [
        _score_color(result.overall_score),
        _COL["score_low"],
        _COL["score_mid"],
        _COL["info_fill"][:6],
    ]
    for i, (lbl, val, col) in enumerate(zip(score_labels, score_values, score_colors), start=1):
        ws.cell(row=row, column=i, value=lbl).font = _font(bold=True, color="718096", size=9)
        ws.cell(row=row, column=i).fill = _fill(_COL["mid_grey"])
        ws.cell(row=row, column=i).alignment = _align(h="center")
        vc = ws.cell(row=row + 1, column=i, value=val)
        vc.font = _font(bold=True, color=col, size=18)
        vc.fill = _fill(_COL["light_grey"])
        vc.alignment = _align(h="center")
        ws.row_dimensions[row + 1].height = 36

    # Category scores table
    row = 9
    _header_row(ws, ["Category", "Score", "Critical", "Warnings", "Passed"], _COL["header_mid"], row)
    ws.row_dimensions[row].height = 22
    for cr in result.category_reports:
        row += 1
        sc = _score_color(cr.score)
        _data_cell(ws, row, 1, cr.name, bold=True)
        _data_cell(ws, row, 2, cr.score, bold=True, color=sc)
        _data_cell(ws, row, 3, cr.critical_count, color=_COL["score_low"] if cr.critical_count else "555555")
        _data_cell(ws, row, 4, cr.warning_count, color=_COL["score_mid"] if cr.warning_count else "555555")
        _data_cell(ws, row, 5, cr.pass_count, color=_COL["score_high"] if cr.pass_count else "555555")
        # Alternating row fill
        if (row % 2) == 0:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = _fill(_COL["mid_grey"])

    ws.freeze_panes = "A4"
    _set_col_widths(ws, [44, 12, 12, 12, 12])


def _sheet_summary_site(wb, cr: "SiteCrawlResult") -> None:
    ws = wb.create_sheet("Site Summary")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"] = "SITE-WIDE SEO AUDIT REPORT"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=14)
    ws["A1"].fill = _fill(_COL["header_dark"])
    ws["A1"].alignment = _align(h="center", v="center", wrap=False)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    ws["A2"] = cr.base_url
    ws["A2"].font = _font(italic=True, color="718096", size=10)
    ws["A2"].fill = _fill(_COL["light_grey"])
    ws["A2"].alignment = _align(h="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = (f"Site type: {cr.profile.label}   |   "
                f"Pages audited: {cr.pages_audited}   |   "
                f"Generated: {datetime.now().strftime('%d %b %Y')}")
    ws["A3"].font = _font(color="718096", size=9)
    ws["A3"].fill = _fill(_COL["mid_grey"])
    ws["A3"].alignment = _align(h="center")

    # Score stats
    row = 5
    for i, (lbl, val, col) in enumerate([
        ("Avg Score", f"{cr.overall_score}/100", _score_color(cr.overall_score)),
        ("Pages Audited", str(cr.pages_audited), "2B6CB0"),
        ("Pages w/ Criticals", str(cr.pages_with_critical), _COL["score_low"]),
        ("Total Critical", str(len(cr.all_critical)), _COL["score_low"]),
        ("Total Warnings", str(len(cr.all_warnings)), _COL["score_mid"]),
        ("Quick Wins", str(len(cr.all_quick_wins)), "2B6CB0"),
    ], start=1):
        ws.cell(row=row, column=i, value=lbl).font = _font(bold=True, color="718096", size=9)
        ws.cell(row=row, column=i).fill = _fill(_COL["mid_grey"])
        ws.cell(row=row, column=i).alignment = _align(h="center")
        vc = ws.cell(row=row + 1, column=i, value=val)
        vc.font = _font(bold=True, color=col, size=16)
        vc.fill = _fill(_COL["light_grey"])
        vc.alignment = _align(h="center")
    ws.row_dimensions[row + 1].height = 36

    # Category avg scores
    row = 9
    _header_row(ws, ["Category", "Avg Score"], _COL["header_mid"], row)
    for cat, avg in cr.category_avg_scores.items():
        row += 1
        sc = _score_color(avg)
        _data_cell(ws, row, 1, cat, bold=True)
        _data_cell(ws, row, 2, avg, bold=True, color=sc)
        if row % 2 == 0:
            for c in range(1, 3):
                ws.cell(row=row, column=c).fill = _fill(_COL["mid_grey"])

    # Top issues across site
    row += 2
    _header_row(ws, ["Most Common Issues Across Site", "Pages Affected"], _COL["header_phase1"], row)
    for check, count in cr.top_issues:
        row += 1
        _data_cell(ws, row, 1, check)
        _data_cell(ws, row, 2, count, bold=True, color=_COL["score_low"])
        if row % 2 == 0:
            for c in range(1, 3):
                ws.cell(row=row, column=c).fill = _fill(_COL["mid_grey"])

    # Per-page scores
    row += 2
    _header_row(ws, ["Page URL", "Score", "Critical", "Warnings", "Quick Wins"], _COL["header_mid"], row)
    for pr in sorted(cr.page_results, key=lambda r: r.overall_score):
        row += 1
        sc = _score_color(pr.overall_score)
        _data_cell(ws, row, 1, pr.url)
        _data_cell(ws, row, 2, pr.overall_score, bold=True, color=sc)
        _data_cell(ws, row, 3, len(pr.critical_findings), color=_COL["score_low"] if pr.critical_findings else "555")
        _data_cell(ws, row, 4, len(pr.warning_findings), color=_COL["score_mid"] if pr.warning_findings else "555")
        _data_cell(ws, row, 5, len(pr.quick_wins), color="2B6CB0" if pr.quick_wins else "555")
        if row % 2 == 0:
            for c in range(1, 6):
                ws.cell(row=row, column=c).fill = _fill(_COL["mid_grey"])

    ws.freeze_panes = "A4"
    _set_col_widths(ws, [55, 12, 14, 14, 12])


def _sheet_findings(wb, findings: list[tuple[str, "Finding"]], sheet_name: str = "All Findings") -> None:
    """
    findings: list of (url, Finding)  — url is "" for single-page mode.
    """
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    headers = ["URL", "Category", "Check", "Severity", "Detail", "Recommendation", "Impact", "Effort"]
    _header_row(ws, headers, _COL["header_mid"])
    ws.row_dimensions[1].height = 20
    ws.auto_filter.ref = f"A1:H1"

    for row_idx, (url, f) in enumerate(findings, start=2):
        sev = f.severity
        bg = _SEV_FILL.get(sev, _COL["white"])
        if f.external_data_needed:
            bg = _COL["ext_fill"]
        cols = [
            url,
            f.category,
            f.check,
            sev.value.upper(),
            f.detail,
            f.recommendation or "",
            f.impact or "",
            f.effort or "",
        ]
        sev_color = {
            Severity.CRITICAL: _COL["score_low"],
            Severity.WARNING:  "B7791F",
            Severity.INFO:     "2B6CB0",
            Severity.PASS:     _COL["score_high"],
        }.get(sev, "2D3748")

        for col_idx, val in enumerate(cols, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = _fill(bg)
            cell.border = _border_thin()
            cell.alignment = _align()
            if col_idx == 4:  # Severity column
                cell.font = _font(bold=True, color=sev_color, size=10)
            else:
                cell.font = _font(size=10)

    ws.freeze_panes = "A2"
    _set_col_widths(ws, [45, 28, 32, 11, 60, 60, 10, 12])


def _sheet_roadmap(wb, result_or_crawl) -> None:
    ws = wb.create_sheet("Roadmap")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"] = "PRIORITISED ACTION ROADMAP"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = _fill(_COL["header_dark"])
    ws["A1"].alignment = _align(h="center", v="center", wrap=False)
    ws.row_dimensions[1].height = 28

    # Collect findings from single or site result
    if hasattr(result_or_crawl, "page_results"):
        # Site crawl
        all_findings = [
            (r.url, f)
            for r in result_or_crawl.page_results
            for f in r.all_findings
        ]
    else:
        all_findings = [("", f) for f in result_or_crawl.all_findings]

    phases = [
        ("Phase 1 — Quick Wins (Week 1–2)",   _COL["header_phase1"],
         [(u, f) for u, f in all_findings
          if f.effort == "Quick Win" and f.severity in (Severity.CRITICAL, Severity.WARNING)]),
        ("Phase 2 — Core Improvements (Month 1–2)", _COL["header_phase2"],
         [(u, f) for u, f in all_findings
          if f.effort == "Medium" and f.severity in (Severity.CRITICAL, Severity.WARNING)]),
        ("Phase 3 — Long-Term Growth (Month 3–6)", _COL["header_phase3"],
         [(u, f) for u, f in all_findings
          if f.effort == "Long-term" and f.severity in (Severity.CRITICAL, Severity.WARNING)]),
    ]

    row = 3
    for phase_title, fill_hex, items in phases:
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = phase_title
        ws[f"A{row}"].font = _font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = _fill(fill_hex)
        ws[f"A{row}"].alignment = _align(h="left", v="center", wrap=False)
        ws.row_dimensions[row].height = 24
        row += 1

        headers = ["Page / Check", "Severity", "Issue", "Recommended Fix"]
        _header_row(ws, headers, _COL["mid_grey"], row)
        for c in range(1, 5):
            ws.cell(row=row, column=c).font = _font(bold=True, color="2D3748", size=10)
        row += 1

        if not items:
            ws.merge_cells(f"A{row}:D{row}")
            ws[f"A{row}"] = "No items in this phase."
            ws[f"A{row}"].font = _font(italic=True, color="718096", size=10)
            row += 1
        else:
            seen_checks: set[str] = set()
            for url, f in items:
                key = f"{f.category}|{f.check}"
                if key in seen_checks and not url:
                    continue  # deduplicate for single-page mode
                seen_checks.add(key)
                page_check = f"{url or ''} | {f.category} — {f.check}".strip(" |")
                bg = _SEV_FILL.get(f.severity, _COL["white"])
                sev_col = {
                    Severity.CRITICAL: _COL["score_low"],
                    Severity.WARNING: "B7791F",
                }.get(f.severity, "2D3748")
                for col_idx, val in enumerate(
                    [page_check, f.severity.value.upper(), f.detail, f.recommendation or ""], start=1
                ):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.fill = _fill(bg)
                    cell.border = _border_thin()
                    cell.alignment = _align()
                    cell.font = _font(
                        bold=(col_idx == 2), size=10,
                        color=(sev_col if col_idx == 2 else "2D3748")
                    )
                row += 1

        row += 1  # gap between phases

    ws.freeze_panes = "A3"
    _set_col_widths(ws, [50, 12, 60, 60])


def _sheet_checklist(wb, result_or_crawl) -> None:
    profile = result_or_crawl.profile
    ws = wb.create_sheet("Site Checklist")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:C1")
    ws["A1"] = f"SITE-TYPE CHECKLIST: {profile.label.upper()}"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = _fill(_COL["header_dark"])
    ws["A1"].alignment = _align(h="center", v="center", wrap=False)
    ws.row_dimensions[1].height = 26

    _header_row(ws, ["Check Item", "Recommendation", "Status"], _COL["header_mid"], 2)
    for row_idx, (name, rec) in enumerate(profile.custom_checks, start=3):
        bg = _COL["light_grey"] if row_idx % 2 == 0 else _COL["white"]
        _data_cell(ws, row_idx, 1, name, fill_hex=bg, bold=True)
        _data_cell(ws, row_idx, 2, rec, fill_hex=bg)
        status_cell = ws.cell(row=row_idx, column=3, value="[ ]")
        status_cell.fill = _fill(bg)
        status_cell.font = _font(size=11)
        status_cell.alignment = _align(h="center")
        status_cell.border = _border_thin()

    ws.freeze_panes = "A3"
    _set_col_widths(ws, [45, 80, 10])


def _sheet_traffic(wb, result_or_crawl) -> None:
    profile = result_or_crawl.profile
    ws = wb.create_sheet("Traffic Strategy")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:B1")
    ws["A1"] = f"TRAFFIC STRATEGY: {profile.label.upper()}"
    ws["A1"].font = _font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill = _fill(_COL["header_green"])
    ws["A1"].alignment = _align(h="center", v="center", wrap=False)
    ws.row_dimensions[1].height = 26

    _header_row(ws, ["#", "Strategy Note"], _COL["header_mid"], 2)
    for i, note in enumerate(profile.traffic_strategy_notes, start=1):
        row = i + 2
        bg = _COL["light_grey"] if i % 2 == 0 else _COL["white"]
        _data_cell(ws, row, 1, i, fill_hex=bg, bold=True, color="2B6CB0")
        _data_cell(ws, row, 2, note, fill_hex=bg)
        ws.row_dimensions[row].height = 36

    # AI section if present
    ai_text = ""
    if hasattr(result_or_crawl, "ai_traffic_strategy"):
        ai_text = result_or_crawl.ai_traffic_strategy
    if ai_text:
        row = len(profile.traffic_strategy_notes) + 4
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = "AI-Generated Traffic Strategy"
        ws[f"A{row}"].font = _font(bold=True, color="FFFFFF", size=11)
        ws[f"A{row}"].fill = _fill(_COL["header_green"])
        ws[f"A{row}"].alignment = _align(h="center")
        row += 1
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = ai_text
        ws[f"A{row}"].font = _font(size=10)
        ws[f"A{row}"].alignment = _align()
        ws[f"A{row}"].fill = _fill(_COL["light_grey"])
        ws.row_dimensions[row].height = max(60, len(ai_text) // 3)

    ws.freeze_panes = "A3"
    _set_col_widths(ws, [6, 110])


# ─── Public API ───────────────────────────────────────────────────────────────

def save_excel(result: "AuditResult", path: str | Path) -> None:
    """Export a single-page AuditResult to Excel."""
    openpyxl = _try_import()[0]
    if openpyxl is None:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _sheet_summary_single(wb, result)

    # All findings
    all_f = [("", f) for f in result.all_findings]
    _sheet_findings(wb, all_f, "All Findings")

    # Quick wins only
    qw_f = [("", f) for f in result.quick_wins]
    if qw_f:
        _sheet_findings(wb, qw_f, "Quick Wins")

    _sheet_roadmap(wb, result)
    _sheet_checklist(wb, result)
    _sheet_traffic(wb, result)

    wb.save(str(path))


def save_site_excel(crawl_result: "SiteCrawlResult", path: str | Path) -> None:
    """Export a full SiteCrawlResult to Excel."""
    openpyxl = _try_import()[0]
    if openpyxl is None:
        raise ImportError("openpyxl is required for Excel export. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _sheet_summary_site(wb, crawl_result)

    # All findings across all pages
    all_f = [
        (r.url, f)
        for r in crawl_result.page_results
        for f in r.all_findings
    ]
    _sheet_findings(wb, all_f, "All Findings")

    # Quick wins across all pages
    qw_f = [
        (r.url, f)
        for r in crawl_result.page_results
        for f in r.quick_wins
    ]
    if qw_f:
        _sheet_findings(wb, qw_f, "Quick Wins")

    _sheet_roadmap(wb, crawl_result)
    _sheet_checklist(wb, crawl_result)
    _sheet_traffic(wb, crawl_result)

    # Per-page detail sheets (only for small crawls)
    if len(crawl_result.page_results) <= 10:
        for pr in crawl_result.page_results:
            # Truncate URL to valid sheet name
            label = pr.url.replace("https://", "").replace("http://", "").replace("/", "_")[:28]
            page_f = [("", f) for f in pr.all_findings]
            _sheet_findings(wb, page_f, label)

    wb.save(str(path))
